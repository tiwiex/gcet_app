import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from gcet_apps.gcet_apps.quotation_margin import export_quotation_margin_analysis as export_quotation_margin
from gcet_apps.gcet_apps.impersonate_user import switch_back
from frappe.utils import get_site_path, fmt_money
import frappe

@frappe.whitelist()
def export_quotation_margin_analysis(quotation_name):
    # Call the function from quotation_export.py
    return export_quotation_margin(quotation_name)


@frappe.whitelist()
def switch_back_impersonator(original_user):
        # call the switch back function
        return switch_back(original_user)



@frappe.whitelist()
def generate_excel(doc_name):
    # Fetch the Quotation document data
    doc = frappe.get_doc("Quotation", doc_name)

    # Create a new Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Quotation Details"

    # Define a thin border for table cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Adjust column widths for better visibility (Shifted by one column)
    column_widths = {
        'B': 5,   # SN (originally A)
        'C': 20,  # Item Name (originally B)
        'D': 30,  # Description (originally C)
        'E': 10,  # Qty (originally D)
        'F': 20,  # Rate (₦) (originally E)
        'G': 20,  # Amount (₦) (originally F)
    }
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    # Header Section: Company Logo and Details (Shifted by one column)
    sheet['B1'] = "Quotation"
    sheet['B1'].font = Font(size=16, bold=True)
    sheet['B1'].alignment = Alignment(horizontal='center')
    
    # Optional: Add company logo if available
    logo_path = get_site_path("public", "files", "Screenshot 2024-10-17 132019.png")
    try:
        img = openpyxl.drawing.image.Image(logo_path)
        img.anchor = 'B2'
        img.width = 150
        sheet.add_image(img)
    except FileNotFoundError:
        pass  # Skip if logo file is not available

    # Customer Information Section (Shifted by one column)
    sheet['B4'] = "Customer Name:"
    sheet['C4'] = doc.customer_name
    sheet['B5'] = "Customer Address:"
    sheet['C5'] = doc.address_display or ""
    sheet['B7'] = "Quotation Date:"
    sheet['C7'] = doc.get_formatted("posting_date")
    sheet['B8'] = "Quotation Number:"
    sheet['C8'] = doc.name
    sheet['B9'] = "Due Date:"
    sheet['C9'] = doc.get_formatted("due_date")

    # Items Table Header (Shifted by one column)
    item_headers = ["SN", "Item Name", "Description", "Qty", "Rate (₦)", "Amount (₦)"]
    sheet.append([''] + item_headers)  # Add an empty cell for column A

    # Style Item Table Header with borders and background color (Shifted by one column)
    for cell in sheet[11]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="90E0EF", end_color="90E0EF", fill_type="solid")
        cell.border = thin_border

    # Populate Item Rows with borders and number format (Shifted by one column)
    for idx, item in enumerate(doc.items, start=1):
        row = [
            '',  # Empty cell for column A
            idx,
            item.item_name,
            item.description or "",
            item.qty,
            item.rate,  # Use numeric value directly
            item.amount  # Use numeric value directly
        ]
        sheet.append(row)

        # Apply right-alignment to numeric cells (Qty, Rate, Amount)
        sheet[f'E{sheet.max_row}'].alignment = Alignment(horizontal='right')
        sheet[f'F{sheet.max_row}'].alignment = Alignment(horizontal='right')
        sheet[f'G{sheet.max_row}'].alignment = Alignment(horizontal='right')

        # Apply number formatting to Rate and Amount
        sheet[f'F{sheet.max_row}'].number_format = '#,##0.00'
        sheet[f'G{sheet.max_row}'].number_format = '#,##0.00'

        for cell in sheet[sheet.max_row]:  # Apply border to each cell in the current row
            cell.border = thin_border

    # Calculate and Display Totals in Column F and G with borders (Shifted by one column)
    current_row = sheet.max_row + 2  # Leave a gap after items

    # Display Total Amount
    sheet[f'F{current_row}'] = "Total Amount:"
    sheet[f'F{current_row}'].font = Font(bold=True)
    sheet[f'F{current_row}'].alignment = Alignment(horizontal='right')
    sheet[f'F{current_row}'].border = thin_border
    sheet[f'G{current_row}'] = doc.net_total  # Use numeric value directly
    sheet[f'G{current_row}'].alignment = Alignment(horizontal='right')
    sheet[f'G{current_row}'].border = thin_border
    sheet[f'G{current_row}'].number_format = '#,##0.00'

    # Display Taxes with borders and number formatting (Shifted by one column)
    if doc.taxes:
        for tax in doc.taxes:
            current_row += 1
            tax_description = "VAT (5%)" if "VAT" in tax.description else tax.description
            sheet[f'F{current_row}'] = tax_description
            sheet[f'F{current_row}'].alignment = Alignment(horizontal='right')
            sheet[f'F{current_row}'].border = thin_border
            sheet[f'G{current_row}'] = tax.tax_amount  # Use numeric value directly
            sheet[f'G{current_row}'].alignment = Alignment(horizontal='right')
            sheet[f'G{current_row}'].border = thin_border
            sheet[f'G{current_row}'].number_format = '#,##0.00'

    # Display Grand Total with borders and number formatting (Shifted by one column)
    current_row += 2  # Leave a gap after taxes
    sheet[f'F{current_row}'] = "Grand Total:"
    sheet[f'F{current_row}'].font = Font(bold=True)
    sheet[f'F{current_row}'].alignment = Alignment(horizontal='right')
    sheet[f'F{current_row}'].border = thin_border
    sheet[f'G{current_row}'] = doc.grand_total  # Use numeric value directly
    sheet[f'G{current_row}'].alignment = Alignment(horizontal='right')
    sheet[f'G{current_row}'].border = thin_border
    sheet[f'G{current_row}'].number_format = '#,##0.00'

    # Additional Notes (Optional) (Shifted by one column)
    current_row += 3
    sheet[f'B{current_row}'] = "Notes:"
    sheet[f'B{current_row}'].border = thin_border
    sheet[f'B{current_row + 1}'] = doc.terms or ""
    sheet[f'B{current_row + 1}'].border = thin_border

    # In Words Section (Shifted by one column)
    sheet[f'B{current_row + 3}'] = "In Words:"
    sheet[f'B{current_row + 3}'].border = thin_border
    sheet[f'B{current_row + 4}'] = doc.in_words or ""
    sheet[f'B{current_row + 4}'].border = thin_border

    # User Signature Section (Shifted by one column)
    current_row += 6
    user_signature = frappe.db.get_value("User Signature", {"user": doc.owner}, "signature")
    user_full_name = frappe.db.get_value("User", {"name": doc.owner}, "full_name")

    if user_signature:
        try:
            signature_img = openpyxl.drawing.image.Image(user_signature)
            signature_img.anchor = f'B{current_row}'
            signature_img.width = 150
            sheet.add_image(signature_img)
        except FileNotFoundError:
            sheet[f'B{current_row}'] = "No signature available"
    else:
        sheet[f'B{current_row}'] = "No signature available"
    
    sheet[f'B{current_row + 2}'] = user_full_name or "Authorized User"

    # Save the Excel file
    file_path = get_site_path("public", "files", f"{doc_name}_quotation.xlsx")
    workbook.save(file_path)

    # Return the relative path for download
    relative_path = f"/files/{doc_name}_quotation.xlsx"
    return relative_path
