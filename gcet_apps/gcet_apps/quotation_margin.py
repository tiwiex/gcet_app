import frappe
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from frappe.utils import get_site_path

@frappe.whitelist()
def export_quotation_margin_analysis(quotation_name):
    # Create a new Excel workbook and worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Quotation Margin Analysis"

    # Define headers for the report
    headers = [
        "Quotation", "Date", "Customer", "Item Name", "Qty", 
        "Total Cost Price", "Total Selling Price", "Margin (%)", "Margin Amount", "Rate"
    ]
    
    # Add headers to the worksheet
    sheet.append(headers)
    
    # Apply styling to the headers
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="90E0EF", end_color="90E0EF", fill_type="solid")

    # Define a thin border for cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Fetch margin analysis data for the specific Quotation
    data = frappe.db.sql("""
        SELECT
            quotation.name AS quotation,
            quotation.transaction_date AS date,
            quotation.customer_name AS customer,
            quotation_item.item_name AS item_name,
            quotation_item.qty AS qty,
            (quotation_item.price_list_rate * quotation_item.qty) AS total_cost_price,
            (quotation_item.rate * quotation_item.qty) AS total_selling_price,
            ROUND(((quotation_item.rate * quotation_item.qty) - (quotation_item.price_list_rate * quotation_item.qty)) / (quotation_item.price_list_rate * quotation_item.qty) * 100, 2) AS margin_percent,
            ((quotation_item.rate * quotation_item.qty) - (quotation_item.price_list_rate * quotation_item.qty)) AS margin_amount,
            quotation_item.rate AS rate
        FROM
            `tabQuotation` AS quotation
        INNER JOIN
            `tabQuotation Item` AS quotation_item
        ON
            quotation.name = quotation_item.parent
        WHERE
            quotation.name = %s
    """, (quotation_name,), as_dict=True)

    # Initialize totals
    total_qty = 0
    total_cost_price = 0
    total_selling_price = 0
    total_margin_amount = 0

    # Populate the worksheet with data
    for idx, row in enumerate(data, start=2):
        sheet.append([
            row['quotation'],
            row['date'],
            row['customer'],
            row['item_name'],
            row['qty'],
            row['total_cost_price'],
            row['total_selling_price'],
            row['margin_percent'],
            row['margin_amount'],
            row['rate']
        ])

        # Sum up totals for the required columns
        total_qty += row['qty']
        total_cost_price += row['total_cost_price']
        total_selling_price += row['total_selling_price']
        total_margin_amount += row['margin_amount']

        # Apply right alignment to numeric cells and borders
        for col in range(5, 11):  # Columns E to K are numeric
            cell = sheet.cell(row=idx, column=col)
            cell.alignment = Alignment(horizontal='right')
            cell.border = thin_border
            # Apply numeric format only to columns with monetary values or percentages
            if col in [5, 6, 7, 8, 9]:  # Total Cost Price, Total Selling Price, Margin (%), Margin Amount, Rate
                cell.number_format = '#,##0.00'

    # Calculate the overall margin percentage
    total_margin_percentage = (
        (total_selling_price - total_cost_price) / total_cost_price * 100
    ) if total_cost_price else 0

    # Add a total row at the end
    total_row = [
        "Total", "", "", "", total_qty, total_cost_price, total_selling_price, 
        total_margin_percentage, total_margin_amount, ""
    ]
    sheet.append(total_row)

    # Style the total row
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='right')
        cell.border = thin_border
        # Only format numeric cells with number_format
        if cell.column in [5, 6, 7, 8, 9]:  # Only numeric cells
            cell.number_format = '#,##0.00'

    # Save the Excel file
    file_path = get_site_path("public", "files", f"{quotation_name}_margin_analysis.xlsx")
    workbook.save(file_path)

    # Return the relative path for download
    return f"/files/{quotation_name}_margin_analysis.xlsx"
